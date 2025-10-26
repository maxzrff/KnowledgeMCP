"""
XLSX document processor.
"""

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.models.document import DocumentFormat
from src.processors.base import BaseProcessor
from src.utils.logging_config import get_logger

logger = get_logger(__name__)


class XLSXProcessor(BaseProcessor):
    """XLSX document processor."""

    @property
    def supported_format(self) -> DocumentFormat:
        return DocumentFormat.XLSX

    async def extract_text(self, file_path: Path) -> str:
        """Extract text from XLSX."""
        try:
            wb = load_workbook(file_path, data_only=True)
            text_parts = []

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text_parts.append(f"Sheet: {sheet_name}")

                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                    if row_text.strip():
                        text_parts.append(row_text)

                text_parts.append("")  # Blank line between sheets

            text = "\n".join(text_parts)
            logger.info(f"Extracted {len(text)} characters from XLSX: {file_path.name}")
            return text
        except Exception as e:
            logger.error(f"Failed to extract text from XLSX {file_path}: {e}")
            raise

    async def extract_metadata(self, file_path: Path) -> dict[str, Any]:
        """Extract metadata from XLSX."""
        try:
            wb = load_workbook(file_path, data_only=True)
            metadata = {
                "sheet_count": len(wb.sheetnames),
                "format": "xlsx",
                "sheets": wb.sheetnames,
            }

            return metadata
        except Exception as e:
            logger.warning(f"Failed to extract XLSX metadata: {e}")
            return {"format": "xlsx"}
